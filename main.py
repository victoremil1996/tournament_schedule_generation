#!/usr/bin/env python3
import pandas as pd
from typing import List, Tuple
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def round_robin_pairings(players: List[str]) -> List[List[Tuple[str, str]]]:
    """Generate single round robin pairings with circle method (even number of players)."""
    n = len(players)
    fixed = players[0]
    others = players[1:]
    rounds = []
    for _ in range(n - 1):
        seq = [fixed] + others
        half = n // 2
        left = seq[:half]
        right = list(reversed(seq[half:]))
        pairs = [(left[i], right[i]) for i in range(half)]
        rounds.append(pairs)
        others = [others[-1]] + others[:-1]
    return rounds

def double_round_robin(single_rounds: List[List[Tuple[str, str]]]) -> List[List[Tuple[str, str]]]:
    """Duplicate rounds with swapped order for a double round robin."""
    mirrored = [[(b, a) for (a, b) in rnd] for rnd in single_rounds]
    return single_rounds + mirrored

def build_schedule_df(players: List[str], rounds: List[List[Tuple[str, str]]]) -> pd.DataFrame:
    tables = len(players) // 2
    data = {}
    for ridx, rnd in enumerate(rounds, start=1):
        cells = [f"{a} vs {b}" for (a, b) in rnd]
        while len(cells) < tables:
            cells.append("")
        data[f"Round {ridx}"] = cells
    df = pd.DataFrame(data)
    df.index = [f"Table {i}" for i in range(1, tables + 1)]
    return df

def build_points_df(players: List[str], num_rounds: int) -> pd.DataFrame:
    return pd.DataFrame(
        "",
        index=[f"Round {i}" for i in range(1, num_rounds + 1)],
        columns=players,
    )

def write_excel(schedule_df: pd.DataFrame, points_df: pd.DataFrame, output_path: str):
    # Write base sheets
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        schedule_df.to_excel(writer, sheet_name="Schedule")
        points_df.to_excel(writer, sheet_name="Points")
    # Add standings with formulas
    wb = load_workbook(output_path)
    ws_points = wb["Points"]
    if "Standings" in wb.sheetnames:
        wb.remove(wb["Standings"])
    ws_stand = wb.create_sheet("Standings")
    ws_stand.append(["Player", "Total Points"])
    for idx, player in enumerate(points_df.columns, start=2):  # players start at col B
        col_letter = ws_points.cell(row=1, column=idx).column_letter
        start_cell = f"{col_letter}2"
        end_cell = f"{col_letter}{points_df.shape[0]+1}"
        total_formula = f"=SUM(Points!{start_cell}:{end_cell})"
        ws_stand.append([player, total_formula])
    # Auto-size columns
    for ws_name in ["Schedule", "Points", "Standings"]:
        ws = wb[ws_name]
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(12, length + 2), 40)
    wb.save(output_path)

def main(players, output="tournament_schedule.xlsx", double_round=False):
    rounds = round_robin_pairings(players)
    if double_round:
        rounds = double_round_robin(rounds)
    schedule_df = build_schedule_df(players, rounds)
    points_df = build_points_df(players, len(rounds))
    write_excel(schedule_df, points_df, output)
    print(f"Created workbook: {output}")
    print(f"Players ({len(players)}): {', '.join(players)}")
    print(f"Rounds: {len(rounds)}; Tables per round: {len(players)//2}")
    print("Sheets: Schedule, Points, Standings.")

if __name__ == "__main__":
    # Define your inputs here:
    players = ["Victor Emil", "Christoffer", "Nis", "Frederik Hansen", "Ask", "Stefan", "William", "Lucas", "Mark", "Mads"]
    output_file = r"tournament_sheets/mtg_2025_09_27.xlsx" #"example_tournament.xlsx"
    double_round = len(players) <= 6  # Set True for double round robin (e.g. with 6 players)

    main(players, output=output_file, double_round=double_round)
